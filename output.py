import csv
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import PROJECT_NAME, SCRIPT_DIR
from output_schemas import build_schema_groups

# BOT8 2.3.0：输出文件使用相对路径 data/YYYYMMDD/HHMMSS.xlsx 与 HHMMSS.csv。
# RUN_START_TIME 在程序启动后固定，确保同一次运行的 Excel 与 CSV 文件名一致。
RUN_START_TIME = datetime.now()
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "data", RUN_START_TIME.strftime("%Y%m%d"))
OUTPUT_XLSX_NAME = os.path.join(OUTPUT_DIR, f"{RUN_START_TIME.strftime('%H%M%S')}.xlsx")
OUTPUT_CSV_NAME = os.path.join(OUTPUT_DIR, f"{RUN_START_TIME.strftime('%H%M%S')}.csv")
OUTPUT_MARKET_CSV_NAME = os.path.join(OUTPUT_DIR, f"{RUN_START_TIME.strftime('%H%M%S')}_market.csv")
OUTPUT_TRADE_FLOW_CSV_NAME = os.path.join(OUTPUT_DIR, f"{RUN_START_TIME.strftime('%H%M%S')}_trade_flow.csv")
OUTPUT_SPLIT_PREFIX = os.path.join(OUTPUT_DIR, f"{RUN_START_TIME.strftime('%H%M%S')}")


def validate_summary_headers(data_log, summary_headers):
    """检查模型实际生成的 summary 字段是否都被输出表头覆盖。"""
    if not data_log:
        return {"missing_headers": [], "unused_headers": []}
    actual_keys = set()
    for row in data_log:
        if isinstance(row, dict):
            actual_keys.update(row.keys())
    header_keys = set(summary_headers)
    return {
        "missing_headers": sorted(actual_keys - header_keys),
        "unused_headers": sorted(header_keys - actual_keys),
    }


def warn_summary_header_mismatch(data_log, summary_headers):
    result = validate_summary_headers(data_log, summary_headers)
    if result["missing_headers"]:
        print("[BOT8 Warning] 以下字段已生成但未写入 SUMMARY_HEADERS：")
        for key in result["missing_headers"]:
            print("  -", key)
    if result["unused_headers"]:
        print("[BOT8 Warning] 以下 SUMMARY_HEADERS 字段当前没有数据：")
        for key in result["unused_headers"]:
            print("  -", key)
    return result

SUMMARY_HEADERS = [
    # 基础状态
    "Turn", "Population", "PopCount", "SharedEnvEnabled",
    "DisasterOccurred", "DisasterType", "DisasterStrength",
    # 环境状态
    "EnvResource", "EnvCapacity", "EnvHealth", "ResourcePressure", "ResourceRegenActual", "EnvConsumption", "EnvHealthChange",
    "EnvResourceUseRate", "ResourceUseToRegenRatio", "LaborResourceUnused", "LaborResourceUnusedRate", "ResourceLimitReached",
    "FoodOperatingStockTarget", "MedicalOperatingStockTarget", "EducationOperatingStockTarget", "ReproductionOperatingStockTarget",
    # 货币与商品库存
    "GovernmentDeposit", "TotalFood", "TotalMedicalGoods", "TotalEducationGoods", "TotalReproductionGoods", "TotalTools",
    "GovernmentFood", "GovernmentMedicalGoods", "GovernmentEducationGoods", "GovernmentReproductionGoods", "GovernmentTools",
    "CompanyMoneyTotal", "CompanyGoodsStockTotal", "CompanyWagesPaid", "CompanyGoodsProduced", "CompanySalesIncome",
    "CompanyInventoryListed", "CompanyInventorySoldToIndividuals", "CompanyInventorySoldToGovernment", "CompanyInventoryUnsold", "CompanyInventoryListingRatio", "CompanyOrderbookAskCount",
    "CompanyHardNeedReleaseEnabledCount", "FoodCompanyHardNeedPressure", "MedicalCompanyHardNeedPressure", "ReproductionCompanyHardNeedPressure",
    "FoodCompanySellableStock", "MedicalCompanySellableStock", "ReproductionCompanySellableStock",
    "FoodCompanyHardNeedReleaseListed", "MedicalCompanyHardNeedReleaseListed", "ReproductionCompanyHardNeedReleaseListed",
    "FoodCompanyInventoryListed", "FoodCompanySoldToIndividuals", "FoodCompanySoldToGovernment", "FoodCompanyInventoryUnsold",
    "MedicalCompanyInventoryListed", "MedicalCompanySoldToIndividuals", "MedicalCompanySoldToGovernment", "MedicalCompanyInventoryUnsold",
    "EducationCompanyInventoryListed", "EducationCompanySoldToIndividuals", "EducationCompanySoldToGovernment", "EducationCompanyInventoryUnsold",
    "ReproductionCompanyInventoryListed", "ReproductionCompanySoldToIndividuals", "ReproductionCompanySoldToGovernment", "ReproductionCompanyInventoryUnsold",
    "NewbornSurvivalSkippedCount", "ReproductionGoodsDemandCount", "ReproductionGoodsDemandBlockedByPoorOldLogic", "ReproductionGoodsDemandBlockedByFood",
    "ReproductionGoodsDemandBlockedBySickOrCritical", "ReproductionGoodsSpendingBlockedByPoorOldLogic",
    "GovernmentPurchaseToCompany",
    "CompanyResourcePurchased", "CompanyResourceCost", "CompanyExpectedProfit", "CompanyActualRevenue", "CompanyWagePaid", "CompanyDividendPaid", "RichTaxIncome", "GovernmentProductionResource",
    "GovernmentOrderbookPurchaseFood", "GovernmentOrderbookPurchaseMedicalGoods", "GovernmentOrderbookPurchaseEducationGoods", "GovernmentOrderbookPurchaseReproductionGoods", "GovernmentOrderbookPurchaseSpending",
    "GovernmentReproductionGoodsReleased", "GovernmentReproductionGoodsReleaseTargets",
    "GovernmentSurplusFoodDeleted", "GovernmentSurplusMedicalGoodsDeleted", "GovernmentSurplusEducationGoodsDeleted", "GovernmentSurplusReproductionGoodsDeleted", "GovernmentSurplusToolsDeleted", "GovernmentSurplusValueTotal",
    "FoodBranchMoney", "FoodBranchStock", "MedicalBranchMoney", "MedicalBranchStock", "EducationBranchMoney", "EducationBranchStock", "ReproductionBranchMoney", "ReproductionBranchStock",
    "CompanyReproductionGoodsStock", "CompanyReproductionGoodsSellable", "ReproductionGoodsHardBuyerCount",
    "ReproductionGoodsHardDemandTotal", "ReproductionGoodsHardDemandSatisfied", "ReproductionGoodsHardDemandUnsatisfied",
    "ReproductionGoodsBlockedNoCompanyStock", "ReproductionGoodsBlockedNoMoney",
    "ReproductionGoodsCompanySalesVolume", "ReproductionGoodsIndividualSalesVolume",
    "IndividualFoodTotal", "IndividualMedicalGoodsTotal", "IndividualEducationGoodsTotal", "IndividualReproductionGoodsTotal", "IndividualToolsTotal",
    "FoodHardShortageCount", "MedicalHardShortageCount", "EducationGoodsShortageCount", "ReproductionGoodsShortageCount",
    "FoodHardDemandTotal", "MedicalHardDemandTotal", "ReproductionHardDemandTotal", "FoodReserveDemandTotal", "MedicalReserveDemandTotal",
    # 生产、市场与政府购买
    "FoodProducedTotal", "MedicalGoodsProducedTotal", "EducationGoodsProducedTotal", "ReproductionGoodsProducedTotal",
    "FoodConsumedTotal", "MedicalGoodsConsumedTotal", "EducationGoodsConsumedTotal", "ReproductionGoodsConsumedTotal",
    "MarketTradeCount", "MarketTradeVolume", "LocalTradeVolume", "ImportVolume", "ExportVolume", "TradeTaxIncome", "ImportTaxIncome", "ImportSpending", "ExportIncome",
    "FoodPriceIndex", "FoodDemand", "FoodSupply", "FoodUnmetDemand", "FoodUnsoldSupply", "FoodLocalTradeVolume", "FoodImportVolume", "FoodExportVolume", "FoodTradeTaxIncome", "FoodImportTaxIncome",
    "MedicalGoodsPriceIndex", "MedicalGoodsDemand", "MedicalGoodsSupply", "MedicalGoodsUnmetDemand", "MedicalGoodsUnsoldSupply", "MedicalGoodsLocalTradeVolume", "MedicalGoodsImportVolume", "MedicalGoodsExportVolume", "MedicalGoodsTradeTaxIncome", "MedicalGoodsImportTaxIncome",
    "EducationGoodsPriceIndex", "EducationGoodsDemand", "EducationGoodsSupply", "EducationGoodsUnmetDemand", "EducationGoodsUnsoldSupply", "EducationGoodsLocalTradeVolume", "EducationGoodsImportVolume", "EducationGoodsExportVolume", "EducationGoodsTradeTaxIncome", "EducationGoodsImportTaxIncome",
    "ReproductionGoodsPriceIndex", "ReproductionGoodsDemand", "ReproductionGoodsSupply", "ReproductionGoodsUnmetDemand", "ReproductionGoodsUnsoldSupply", "ReproductionGoodsLocalTradeVolume", "ReproductionGoodsImportVolume", "ReproductionGoodsExportVolume", "ReproductionGoodsTradeTaxIncome", "ReproductionGoodsImportTaxIncome",
    "GovernmentPurchaseFood", "GovernmentPurchaseMedicalGoods", "GovernmentPurchaseSpending",
    "GovernmentStockpileFood", "GovernmentStockpileMedicalGoods", "GovernmentStockpileSpending",
    "GovernmentReleaseFood", "GovernmentReleaseMedicalGoods", "GovernmentReleaseIncome", "GovernmentSubsidyValue", "MarketStabilityIndex",
    # 生存、医疗、阶层与信任
    "FoodShortageCount", "MedicalShortageCount", "SickCount", "NewSickCount",
    "Security", "MedicalLevel", "Trust", "TrustChange", "CriticalCount",
    "PoorCount", "LowerCount", "MiddleCount", "RichCount", "UpwardMobilityCount", "DownwardMobilityCount", "SameClassCount", "UpwardMobilityRate", "DownwardMobilityRate",
    # 财富与能力
    "TotalBalance", "AvgBalance", "MedianBalance", "MinBalance", "MaxBalance",
    "AvgIntelligence", "AvgStrength", "AvgMorality", "AvgReproduce", "AvgLabor",
    "AvgHealthIndex", "AvgEducationCapital", "AvgReproductiveSecurity", "MedicalRecoveryCount",
    "HealthDeteriorationCount", "AvgReproductiveSecurityBonus", "CompanyResilienceScore", "GovernmentPolicyPressureScore",
    "AvgDeadLifeSpan", "Gini",
    # 劳动与税收
    "LaborCandidateCount", "LaborWorkerCount", "WorkersPaidCount", "WorkersPaidEnoughForFoodCount", "WorkersPaidEnoughForReproductionCount", "LaborRequestTotal", "LaborAllocatedTotal", "LaborUnmetDemand", "AvgAllocatedResource",
    "ProductionBudgetTotal", "GoodsProducedTotal", "GoodsTaxTotal", "WealthTaxTotal",
    # 风险行为与公共机制
    "InternalPlunderCount", "InternalPlunderTotalValueLoss", "InternalPlunderTotalValueGain", "SanctionCount",
    "InvasionAttemptCount", "InvasionSuccessCount", "InvasionTotalValueLoss", "InvasionTotalValueGain",
    "GovernmentAidTotal", "GovernmentEducationTotal", "GovAidBudgetUsed", "GovAidBudgetRemaining", "IndividualRescueTotal", "MoralDonationTotal", "MoralDonationCount",
    # 人口与演化
    "SurvivalCostTotal", "ReproductionGoodsConsumed", "ChildInitialWealthTotal", "InheritanceTransferTotal",
    "BirthCount", "ReproductionEligibleCount", "ReproductionAttemptCount",
    "SecondaryBirthEligibleCount", "SecondaryBirthConditionReadyCount", "SecondaryBirthAttemptCount", "SecondaryBirthSuccessCount",
    "SecondaryBirthBlockedSickOrCritical", "SecondaryBirthBlockedNoReproductionGoods", "SecondaryBirthBlockedNoFoodSafety", "SecondaryBirthBlockedLowReproduceChance",
    "BirthBlockedCritical", "BirthBlockedSick", "BirthBlockedNoMoney", "BirthBlockedNoReproductionGoods", "BirthBlockedNoFoodSafety", "BirthBlockedLowReproduceChance", "BirthBlockedOther", "CumulativeBirthCount", "CumulativeDeathCount", "BirthDeathRatio", "CumulativeBirthBlockedCritical", "CumulativeBirthBlockedSick", "CumulativeBirthBlockedNoMoney", "CumulativeBirthBlockedNoReproductionGoods", "CumulativeBirthBlockedNoFoodSafety", "CumulativeBirthBlockedLowReproduceChance", "CumulativeBirthBlockedOther", "BirthFoodTransferredToChild", "DeathCount", "EnteredCriticalCount", "RecoveredFromCriticalCount",
    "SingleSurvivorTurnCount", "TurnsAtPopulationBelow3", "LastSurvivorDeathReason", "LastSurvivorReproduceChanceFailed", "LastSurvivorHadReproductionGoods", "LastSurvivorHadFoodForBirth",
    "AvgAgeRound", "MinAgeRound", "MaxAgeRound", "AvgLifeRemaining", "MinLifeRemaining", "MaxLifeRemaining",
    "DeathsByLifeEndWhenPopulationBelow3", "DeathsByLifeEndWhenPopulationBelow5",
    "DeathsByFoodShortageWhenPopulationBelow5", "DeathsByMedicalShortageWhenPopulationBelow5", "DeathsByCriticalGoodsShortageWhenPopulationBelow5",
    "EnteredCriticalWhenPopulationBelow5", "RecoveredCriticalWhenPopulationBelow5",
    "WorkersWhenPopulationBelow5", "WagePaidWhenPopulationBelow5", "FoodBoughtWhenPopulationBelow5", "CompanyFoodStockWhenPopulationBelow5", "GovernmentFoodWhenPopulationBelow5",
    "FoodGini", "MoneyGini", "MedicalGoodsGini", "ReproductionGoodsGini",
    "Bottom20FoodAvg", "Bottom20MoneyAvg", "Bottom20MedicalGoodsAvg", "Bottom20ReproductionGoodsAvg",
    "FoodBelowSurvivalCostCount", "FoodZeroCount", "MoneyZeroCount", "MedicalGoodsZeroCount", "ReproductionGoodsZeroCount",
    "FoodAidEligibleCount", "FoodAidReceivedCount", "FoodAidUnmetCount", "GovernmentFoodBeforeAid", "GovernmentFoodAfterAid", "FoodShortageWithGovernmentFoodCount",
    "MedicalAidEligibleCount", "MedicalAidReceivedCount", "MedicalAidUnmetCount", "GovernmentMedicalGoodsBeforeAid", "GovernmentMedicalGoodsAfterAid", "MedicalShortageWithGovernmentMedicalGoodsCount",
    "CriticalMedicalNeedCount", "CriticalMedicalAidReceivedCount", "CriticalMedicalAidUnmetCount", "CriticalRecoveredCount", "CriticalDiedCount",
    "MedicalGoodsBoughtByCritical", "MedicalGoodsBoughtByHealthy", "CompanyMedicalGoodsStock",
    "Age0Count", "Age1To3Count", "Age4To8Count", "Age0AvgFood", "Age0AvgMoney", "Age0AvgMedicalGoods", "Age0AvgReproductionGoods", "Age0CriticalCount",
    "Age1To3AvgFood", "Age1To3AvgMoney", "Age1To3AvgMedicalGoods", "Age1To3AvgReproductionGoods", "Age1To3CriticalCount",
    "TotalWagesPaid", "AvgWagePerWorker", "MedianWagePerWorker", "MinWagePerWorker", "WorkersCount",
    "WageToSurvivalCostRatio", "WageToFoodPriceRatio", "WageToReproductionGoodsPriceRatio", "Bottom20WageAvg",
    "CompanyCashBeforeWages", "CompanyCashAfterWages", "CompanyCashAfterResourcePurchase", "CompanyUnableToPayFullWagesCount", "CompanyProductionStoppedByCashCount", "CompanyProductionStoppedByStockCount",
    "EffectiveBuyWillingnessAvg", "WageConsumptionBonusAvg", "WageResponsiveBuyerCount", "WageResponsiveExtraCapTotal",
    "WageFundedMarketSpending", "WorkerMarketSpending", "WorkerMarketSpendingToCompany",
    "FoodHardNeedCount", "FoodHardNeedAmount", "FoodHardSpendingCap", "FoodHardActualSpending", "FoodHardSatisfiedAmount", "FoodHardUnsatisfiedAmount",
    "MedicalHardNeedCount", "MedicalHardNeedAmount", "MedicalHardSpendingCap", "MedicalHardActualSpending", "MedicalHardSatisfiedAmount", "MedicalHardUnsatisfiedAmount",
    "ReproductionHardNeedCount", "ReproductionHardNeedAmount", "ReproductionHardSpendingCap", "ReproductionHardActualSpending", "ReproductionHardSatisfiedAmount", "ReproductionHardUnsatisfiedAmount",
    "HardNeedSpendingTotal", "ReserveNeedSpendingTotal",
    "HardNeedBlockedByNoCash", "HardNeedBlockedByNoMarketStock", "HardNeedBlockedByHighPrice", "HardNeedBlockedByBudgetCap",
    "FoodHardNeedSatisfiedRate", "MedicalHardNeedSatisfiedRate", "ReproductionHardNeedSatisfiedRate",
    "InventorySalesDividendPaid", "InventorySalesDividendRecipients", "HistoricalInventorySalesIncome",
    "InventorySalesDividendEligibleBranches", "InventorySalesDividendBlockedByCashProtection",
    "InventorySalesDividendBlockedByNoHistoricalIncome", "InventorySalesDividendCashFloor",
    "ExcessCashDividendPaid", "ExcessCashDividendRecipients", "ExcessCashDividendPool", "ExcessCashDividendEligibleBranches",
    "ExcessCashDividendBlockedByNoExcessCash", "ExcessCashDividendBlockedByNoRecipients",
    "LaborWorkerCountWhenPopBelow5", "TurnsWithNoWorkersWhenPopBelow5", "TurnsWithNoWagesWhenPopBelow5",
    "CompanyHasCashButNoWorkersCount", "CompanyHasStockButNoWorkersCount",
    "PopBelow5TurnCount", "PopBelow3TurnCount",
    "LaborEligibleCountWhenPopBelow5", "LaborEligibleCountWhenPopBelow3",
    "LaborWillingCountWhenPopBelow5", "LaborWillingCountWhenPopBelow3",
    "ActualWorkerCountWhenPopBelow5", "ActualWorkerCountWhenPopBelow3",
    "NoWorkerReasonSickCount", "NoWorkerReasonCriticalCount", "NoWorkerReasonLowLaborCount",
    "NoWorkerReasonNoCompanyDemandCount", "NoWorkerReasonNoExpectedProfitCount", "NoWorkerReasonNoResourceCount",
    "CompanyDemandForWorkersWhenPopBelow5", "CompanyDemandForWorkersWhenPopBelow3",
    "BranchesWithPositiveExpectedProfitWhenPopBelow5", "BranchesWithPositiveExpectedProfitWhenPopBelow3",
    "BranchesStoppedByStockWhenPopBelow5", "BranchesStoppedByCashWhenPopBelow5", "BranchesStoppedByResourceWhenPopBelow5",
    "GovernmentProductionResourceWhenPopBelow5", "CompanyTotalStockWhenPopBelow5", "CompanyTotalMoneyWhenPopBelow5",
    "LowPopSnapshotCount", "LastIndividualsAvgMoney", "LastIndividualsAvgFood",
    "LastIndividualsAvgMedicalGoods", "LastIndividualsAvgReproductionGoods",
    "LastIndividualsAvgLabor", "LastIndividualsAvgReproduce", "LastIndividualsAvgLifeRemaining", "LastIndividualsAvgAge",
    "LastIndividualsSickCount", "LastIndividualsCriticalCount", "LastIndividualsCanWorkCount", "LastIndividualsCanReproduceCount",
    "LastIndividualsHasFoodForBirthCount", "LastIndividualsHasReproductionGoodsCount",
    "LifeEndWithCanReproduce", "LifeEndWithCanWork", "LifeEndWithFoodAndReproductionGoods",
    "LastDeathLifeRemaining", "LastDeathHadFoodForBirth", "LastDeathHadReproductionGoods",
    "LastDeathWasSick", "LastDeathWasCritical", "LastDeathCouldWork", "LastDeathCouldReproduce",
    "Last3PopulationAvgLifeRemaining", "Last3PopulationMinLifeRemaining", "Last3PopulationReproductiveEligibleCount",
    "DeathByLifeEndWithReproductionGoods", "DeathByLifeEndWithFoodForBirth", "TurnDeathByLifeEndWithReproductionGoods", "TurnDeathByLifeEndWithFoodForBirth",
    "LaborCandidateRawCount", "LaborCandidatesTrimmedByTendency", "LaborAllocatedCandidateCount", "LaborCandidatesWithoutAllocation",
    "LaborPositiveProfitButNoWorkerWhenPopBelow5",
    "ParentFoodRequirement", "PotentialParentCountWhenPopBelow5", "PotentialParentWithReproductionGoodsWhenPopBelow5",
    "PotentialParentFoodReadyWhenPopBelow5", "ParentFoodGapWhenPopBelow5", "ParentFoodGapWhenPopBelow3", "LastIndividualsParentFoodGapAvg",
    "FoodBoughtByPotentialParent", "FoodAidToPotentialParent", "BirthBlockedFoodSafetyWithReproductionGoods",
    "EvolutionMorality", "EvolutionStrength", "EvolutionReproduce", "EvolutionLabor", "EvolutionReady",
    "EvolutionSampleCount", "EvolutionDirectionChangeCount", "EvolutionFitnessAvg",
    "EvolutionFitnessGapMorality", "EvolutionFitnessGapStrength", "EvolutionFitnessGapReproduce", "EvolutionFitnessGapLabor",
    "EvolutionSignalMorality", "EvolutionSignalStrength", "EvolutionSignalReproduce", "EvolutionSignalLabor",
    # dev44：补齐已生成但旧 SUMMARY_HEADERS 未覆盖的字段。
    "UsePopulationScaledInitials",
    "CompanyInitialMoneyEffective",
    "GovernmentInitialMoneyEffective",
    "CompanyInitialEducationStockTarget",
    "CompanyInitialReproductionStockTarget",
    "EducationCompanySellableStock",
    "EducationInventoryResilienceGap",
    "EducationInventoryResilienceWeightAdded",
    "ReproductionInventoryResilienceGap",
    "ReproductionInventoryResilienceWeightAdded",
    "HardNeedProductionResponseEnabled",
    "FoodHardNeedProductionBonus",
    "MedicalHardNeedProductionBonus",
    "ReproductionHardNeedProductionBonus",
    "EducationNeedProductionBonus",
    "FoodHardNeedUnmetForProduction",
    "MedicalHardNeedUnmetForProduction",
    "ReproductionHardNeedUnmetForProduction",
    "EducationNeedUnmetForProduction",
    "ToolsProducedTotal",
    "ToolsConsumedTotal",
    "MarketFoodVolume",
    "MarketMedicalGoodsVolume",
    "MarketEducationGoodsVolume",
    "MarketReproductionGoodsVolume",
    "PopulationResourceClaim",
    "PopulationResourceQuota",
    "PopulationResourceUsed",
    "PopulationResourceShortage",
    "InternalPlunderVictimLoss",
    "InternalPlunderGain",
    "InternalPlunderSystemLoss",
    "InvasionVictimLoss",
    "InvasionGainTotal",
    "InvasionSystemLoss",
]

INDIVIDUAL_HEADERS = [
    # 身份、阶层与状态
    "Turn", "Population", "ID", "Code", "AncestorCode", "AncestorIndex", "LineageSequence", "BirthTurn", "ParentClass", "BirthClass", "CurrentClass", "ClassChange", "IsUpwardMobile", "IsDownwardMobile",
    "AgeRound", "InitialAgeRounds", "Life", "SurvivalRounds", "Critical", "UsedCriticalChance", "CharityBanned", "Role", "TribeTrust",
    # 个体属性
    "Morality", "Strength", "Intelligence", "TemporaryIntelligence", "EffectiveIntelligence", "AbilityTotal", "Reproduce", "Labor",
    # 核心库存
    "TurnStartBalance", "EndBalance", "Balance", "Money", "TurnStartMarketValue", "PreSurvivalMarketValue", "Food", "MedicalGoods", "EducationGoods", "ReproductionGoods", "Tools",
    # 市场与劳动核心结果
    "DidMarketTrade", "MarketMoneySpent", "MarketMoneyEarned", "MarketImportValue", "MarketExportValue", "MarketTaxPaid",
    "MarketUnmetFoodNeed", "MarketUnmetMedicalNeed", "MarketUnmetEducationNeed", "MarketUnmetReproductionNeed",
    "TurnIncome", "DidLabor", "EmployerBranch", "WageReceived", "DividendReceived", "ProducedGoodsValue", "PrimaryProductionGood", "ProductionPriceResponse", "TotalMarketNeed", "TotalMarketUnmetNeed", "LaborParticipationChance", "RequestedResource", "AllocatedResource", "EnvConsumedByLabor", "LaborNetProduction",
    "FoodProduced", "MedicalGoodsProduced", "EducationGoodsProduced", "ReproductionGoodsProduced",
    "FoodConsumed", "MedicalGoodsConsumed", "EducationGoodsConsumed", "ReproductionGoodsConsumed",
    "FoodTaxPaid", "MedicalGoodsTaxPaid", "EducationGoodsTaxPaid", "ReproductionGoodsTaxPaid",
    # 生存、医疗、教育、生育
    "FoodAidReceived", "MedicalAidReceived", "IsSick", "BecameSickThisTurn", "SicknessRisk",
    "HealthIndex", "HealthDeltaThisTurn", "MedicalRecoveryThisTurn", "HealthDeterioratedThisTurn",
    "EducationCapital", "ReproductiveSecurityScore", "ReproductiveSecurityBonus",
    "MedicalGoodsNeeded", "MedicalGoodsShortage", "FoodShortage",
    "EducationGoodsUsedForChild", "ReproductionGoodsUsed", "MoneyUsedForReproduction",
    "WealthTaxPaid", "RichTaxPaid", "TotalTaxPaid",
    # 掠夺、侵略、救助、繁殖、生存
    "DidInternalPlunder", "WasInternalPlunderVictim", "WasSanctioned", "SanctionLoss", "InternalPlunderTotalValueGain", "InternalPlunderTotalValueLoss",
    "DidInvasion", "InvasionSuccess", "WasInvasionVictim", "InvasionTotalValueGain", "InvasionTotalValueLoss", "InvasionFailLifeLoss",
    "GovernmentAidReceived", "IndividualRescueGiven", "IndividualRescueReceived", "MoralDonationGiven", "MoralDonationReceived",
    "DidReproduce", "ChildCount", "ReproductionGoodsConsumed", "ReproductionMoneyTransferredToChild", "BirthFoodTransferredToChild", "BirthFoodReceived", "InheritanceGiven", "InheritanceReceived",
    "EducationTempIntelligenceReceived", "GovernmentEducationInvestmentReceived", "GovernmentEducationTempIntelligenceReceived", "EducationTempIntelligenceGiven",
    "SurvivalCostPaid", "EnteredCriticalThisTurn", "RecoveredFromCriticalThisTurn", "DiedThisTurn", "DeathReason", "DepositToGovernmentOnDeath",
]


def ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def save_excel_summary(summary_rows):
    if not summary_rows:
        return
    ensure_output_dir()
    wb = Workbook()
    ws = wb.active
    ws.title = "PopulationSummary"
    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"] = f"{PROJECT_NAME} - Population Summary"
    ws["A1"].font = white_font
    ws["A1"].fill = title_fill
    for c, h in enumerate(SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    for r, row in enumerate(summary_rows, start=3):
        for c, h in enumerate(SUMMARY_HEADERS, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(h, ""))
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
    ws.freeze_panes = "A3"
    for c in range(1, len(SUMMARY_HEADERS) + 1):
        letter = get_column_letter(c)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in ws[letter])
        ws.column_dimensions[letter].width = min(max_len + 2, 24)
    wb.save(OUTPUT_XLSX_NAME)


def save_individual_csv(individual_rows):
    # 即使本次运行尚未达到个体记录间隔，也创建带表头的 CSV，避免用户误以为未正常输出。
    ensure_output_dir()
    with open(OUTPUT_CSV_NAME, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=INDIVIDUAL_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(individual_rows)


MARKET_HEADERS = [
    "Turn", "Population", "Goods", "PriceIndex", "Demand", "Supply", "TradeVolume", "UnmetDemand", "UnsoldSupply",
    "LocalTradeVolume", "ImportVolume", "ExportVolume", "TradeTaxIncome", "ImportTaxIncome",
    "CompanyInventoryListed", "CompanySoldToIndividuals", "CompanySoldToGovernment", "CompanyInventoryUnsold",
    "GovernmentOrderbookPurchase", "GovernmentSurplusDeleted", "EffectiveUnmetDemand", "EffectiveUnsoldSupply", "MarketStabilityIndex"
]

def build_market_rows(summary_rows):
    rows = []
    goods_map = [
        ("Food", "Food", "Food"),
        ("MedicalGoods", "MedicalGoods", "Medical"),
        ("EducationGoods", "EducationGoods", "Education"),
        ("ReproductionGoods", "ReproductionGoods", "Reproduction"),
    ]
    for row in summary_rows:
        for label, prefix, company_prefix in goods_map:
            rows.append({
                "Turn": row.get("Turn", ""),
                "Population": row.get("Population", ""),
                "Goods": label,
                "PriceIndex": row.get(f"{prefix}PriceIndex", ""),
                "Demand": row.get(f"{prefix}Demand", ""),
                "Supply": row.get(f"{prefix}Supply", ""),
                "TradeVolume": row.get(f"{prefix}TradeVolume", row.get(f"Market{prefix}Volume", "")),
                "UnmetDemand": row.get(f"{prefix}UnmetDemand", ""),
                "UnsoldSupply": row.get(f"{prefix}UnsoldSupply", ""),
                "LocalTradeVolume": row.get(f"{prefix}LocalTradeVolume", ""),
                "ImportVolume": row.get(f"{prefix}ImportVolume", ""),
                "ExportVolume": row.get(f"{prefix}ExportVolume", ""),
                "TradeTaxIncome": row.get(f"{prefix}TradeTaxIncome", ""),
                "ImportTaxIncome": row.get(f"{prefix}ImportTaxIncome", ""),
                "CompanyInventoryListed": row.get(f"{company_prefix}CompanyInventoryListed", ""),
                "CompanySoldToIndividuals": row.get(f"{company_prefix}CompanySoldToIndividuals", ""),
                "CompanySoldToGovernment": row.get(f"{company_prefix}CompanySoldToGovernment", ""),
                "CompanyInventoryUnsold": row.get(f"{company_prefix}CompanyInventoryUnsold", ""),
                "GovernmentOrderbookPurchase": row.get(f"GovernmentOrderbookPurchase{prefix}", ""),
                "GovernmentSurplusDeleted": row.get(f"GovernmentSurplus{prefix}Deleted", ""),
                "EffectiveUnmetDemand": row.get(f"{prefix}UnmetDemand", ""),
                "EffectiveUnsoldSupply": row.get(f"{prefix}UnsoldSupply", ""),
                "MarketStabilityIndex": row.get("MarketStabilityIndex", ""),
            })
    return rows

def save_market_csv(summary_rows):
    market_rows = build_market_rows(summary_rows)
    if not market_rows:
        return
    ensure_output_dir()
    with open(OUTPUT_MARKET_CSV_NAME, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=MARKET_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(market_rows)


TRADE_FLOW_HEADERS = [
    "Turn", "Goods", "BuyerTribe", "SellerTribe", "Amount", "SellerPriceIndex", "GoodsValue", "TradeTax", "ImportTax", "TotalPaid", "BuyerID", "SellerID"
]


def save_split_summary_csvs(summary_rows):
    """Save categorized summary CSV files using output_schemas groups.

    This is a non-breaking system-stage output split: the legacy Excel/CSV outputs are
    still produced, while these additional files make large runs easier to inspect.
    """
    if not summary_rows:
        return
    ensure_output_dir()
    schema_groups = build_schema_groups(SUMMARY_HEADERS)
    file_suffixes = {
        "summary_core": "summary_core",
        "market": "market_summary",
        "company": "company",
        "government": "government",
        "diagnostics": "diagnostics",
    }
    for group_name, fields in schema_groups.items():
        if not fields:
            continue
        path = f"{OUTPUT_SPLIT_PREFIX}_{file_suffixes[group_name]}.csv"
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(summary_rows)

def save_trade_flow_csv(trade_flow_rows):
    ensure_output_dir()
    with open(OUTPUT_TRADE_FLOW_CSV_NAME, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=TRADE_FLOW_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(trade_flow_rows)

def save_outputs(env):
    warn_summary_header_mismatch(env.summary_output_rows, SUMMARY_HEADERS)
    save_excel_summary(env.summary_output_rows)
    save_individual_csv(env.individual_output_rows)
    save_market_csv(env.summary_output_rows)
    save_split_summary_csvs(env.summary_output_rows)
    save_trade_flow_csv(getattr(env, "trade_flow_rows", []))
